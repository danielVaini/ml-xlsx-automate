from tkinter import *
from tkinter import filedialog
import ttkbootstrap as tb
import openpyxl
from ttkbootstrap.dialogs import Messagebox
import random


root = tb.Window(themename="superhero")

root.title("MP Automate")
root.geometry('800x800')

df_models = False
df_anuncio = False
# Criar método para carregar planilha Modelos/MP
def open_file():
    global df_anuncio
    filename = filedialog.askopenfilename()
    if filename:
        df = openpyxl.load_workbook(filename)
        df_anuncio = df

def open_file_models():
    global df_models
    filename = filedialog.askopenfilename()
    if filename:
        df = openpyxl.load_workbook(filename)
        df_models = df

# Criar métodos para preencher planilha com os dados
def gerar_numero_aleatorio(minimo, maximo, titulos):
    # Garantir que o valor minimo e máximo estejam ao menos uma vez
    numeros = [minimo, maximo]
    for _ in range(len(titulos) - 2):
        numeros.append(random.uniform(minimo, maximo))

    return numeros

def generate_title(title, models):
    titles = []
    for m in models:
        aux = title.replace('_', str(m))
        titles.append(aux)
    return titles

def add_data():
    if not df_models or not df_anuncio:

        if not df_models: Messagebox.ok("Planilha de modelos não carregada", title="Planilha modelos", alert=True) 
        if not df_anuncio: Messagebox.ok("Planilha de anuncio não carregada", title="Planilha anuncio", alert=True) 

    elif not title.get() or not estoque.get() or not preco.get() or not marca.get() or not descricao.get():

        if not title.get(): Messagebox.ok("Titulo em branco", title="Campo em branco", alert=True) 
        if not estoque.get(): Messagebox.ok("Estoque em branco", title="Campo em branco", alert=True) 
        if not preco.get(): Messagebox.ok("Preço em branco", title="Campo em branco", alert=True) 
        if not marca.get(): Messagebox.ok("Marca em branco", title="Campo em branco", alert=True) 
        if not descricao.get(): Messagebox.ok("Descricao em branco", title="Campo em branco", alert=True) 
        
    elif condicao.get() == "Selecione uma opção" or frete.get() == "Selecione uma opção" or tipo_anuncio.get() == "Selecione uma opção" or forma_envio.get() == "Selecione uma opção" or retirada.get() == "Selecione uma opção":

        if condicao.get() == "Selecione uma opção": Messagebox.ok("Condição em branco", title="Campo em branco", alert=True) 
        if frete.get() == "Selecione uma opção": Messagebox.ok("Frete em branco", title="Campo em branco", alert=True) 
        if tipo_anuncio.get() == "Selecione uma opção": Messagebox.ok("Tipo de anuncio em branco", title="Campo em branco", alert=True) 
        if forma_envio.get() == "Selecione uma opção": Messagebox.ok("Forma de envio em branco", title="Campo em branco", alert=True) 
        if retirada.get() == "Selecione uma opção": Messagebox.ok("Forma de retirada em branco", title="Campo em branco", alert=True) 

    else:
        aba_outros = df_anuncio['Outros']

        nomes_colunas = [coluna.value for coluna in aba_outros[3]]

        models = df_models.active  
        modelos = [cell.value for cell in models['A']]
        titles = generate_title(title.get(), modelos)

        linha_requerida = int(linha.get()) if linha.get() else 6

        if preco_max.get():
            precos = gerar_numero_aleatorio(int(preco.get()), int(preco_max.get()), titles)
        else:
            precos = preco.get()


        for i, e in enumerate(titles, start=1):

            if varia_frete.get():
                if i % 2 == 0:
                    frete_value = "Por conta do comprador"
                else:
                    frete_value = "Você oferece frete grátis"
            else:
                frete_value = frete.get()

                
            # Encontre o índice da coluna desejada
            indice_coluna_title                 = nomes_colunas.index('Título: informe o produto, marca, modelo e destaque as características principais') + 1  # Adicione 1 porque os índices começam em 1 no openpyxl
            indice_coluna_estoque               = nomes_colunas.index('Estoque') + 1  
            indice_coluna_preco                 = nomes_colunas.index('Preço [R$]') + 1  
            indice_coluna_condicao              = nomes_colunas.index('\u200bCondição') + 1  
            indice_coluna_descricao             = nomes_colunas.index('Descrição') + 1  
            indice_coluna_tipo_anuncio          = nomes_colunas.index('Tipo de anúncio') + 1  
            indice_coluna_forma_envio           = nomes_colunas.index('Forma de envio') + 1  
            indice_coluna_frete                 = nomes_colunas.index('Frete') + 1  
            indice_coluna_retirada_pessoalmente = nomes_colunas.index('Retirar pessoalmente') + 1  
            indice_coluna_marca                 = nomes_colunas.index('Marca') + 1  

            # Acesse os dados nas linhas subsequentes e altere um valor específico (por exemplo, linha 4, coluna 'Nome')
            linha_alvo = linha_requerida + i
            aba_outros.cell(row=linha_alvo, column=indice_coluna_title, value=e)
            aba_outros.cell(row=linha_alvo, column=indice_coluna_estoque, value=estoque.get())

            if len(precos) > 1:
                aba_outros.cell(row=linha_alvo, column=indice_coluna_preco, value=round(float(precos[i-1]), 2))
            else:
                aba_outros.cell(row=linha_alvo, column=indice_coluna_preco, value=(float(precos), 2))

            aba_outros.cell(row=linha_alvo, column=indice_coluna_condicao, value=condicao.get())
            aba_outros.cell(row=linha_alvo, column=indice_coluna_tipo_anuncio, value=tipo_anuncio.get())
            aba_outros.cell(row=linha_alvo, column=indice_coluna_forma_envio, value=forma_envio.get())
            aba_outros.cell(row=linha_alvo, column=indice_coluna_frete, value=frete_value)
            aba_outros.cell(row=linha_alvo, column=indice_coluna_retirada_pessoalmente, value=retirada.get())
            aba_outros.cell(row=linha_alvo, column=indice_coluna_marca, value=marca.get())
            aba_outros.cell(row=linha_alvo, column=indice_coluna_descricao, value=descricao.get())


def save_file():
    filename = filedialog.asksaveasfilename(defaultextension=".xlsx")
    if filename:
        df_anuncio.save(filename)

# Criar método para salvar planilha preenchida

# Criar botões para entrada de dados
title = tb.Entry(root, bootstyle="primary")
t_label = tb.Label(root, bootstyle="primary", text="Titulo")

title.grid(row=2, column=1, padx=10, pady=10)
t_label.grid(row=2, column=0, padx=10, pady=10)


estoque = tb.Entry(root, bootstyle="primary")
es_label = tb.Label(root, bootstyle="primary", text="Estoque")

estoque.grid(row=3, column=1, padx=10, pady=10)
es_label.grid(row=3, column=0, padx=10, pady=10)

linha = tb.Entry(root, bootstyle="primary")
row_label = tb.Label(root, bootstyle="primary", text="Iniciar na linha")

linha.grid(row=3, column=3, padx=10, pady=10)
row_label.grid(row=3, column=2, padx=10, pady=10)


preco = tb.Entry(root, bootstyle="primary")
pr_label = tb.Label(root, bootstyle="primary", text="Preço")

preco_max = tb.Entry(root, bootstyle="primary")
pr_min_label  = tb.Label(root, bootstyle="primary", text="Preço max")

preco.grid(row=4, column=1, padx=10, pady=10)
pr_label.grid(row=4, column=0, padx=10, pady=10)

pr_min_label.grid(row=4, column=2, padx=10, pady=10)
preco_max.grid(row=4, column=3, padx=10, pady=10)

marca = tb.Entry(root, bootstyle="primary")
mr_label = tb.Label(root, bootstyle="primary", text="Marca")

marca.grid(row=5, column=1, padx=10, pady=10)
mr_label.grid(row=5, column=0, padx=10, pady=10)

descricao = tb.Entry(root, bootstyle="primary")
dr_label = tb.Label(root, bootstyle="primary", text="Descrição")

descricao.grid(row=6, column=1, padx=10, pady=10)
dr_label.grid(row=6, column=0, padx=10, pady=10)

link = tb.Entry(root, bootstyle="primary")
link_label = tb.Label(root, bootstyle="primary", text="Link fotos")

link.grid(row=7, column=1, padx=10, pady=10)
link_label.grid(row=7, column=0, padx=10, pady=10)


# Condição produto
select_opt = ['Novo', 'Usado', 'Recondicionado']
condicao = tb.Combobox(root, bootstyle="primary", values=select_opt)
condicao.set("Selecione uma opção")

tb.Label(root, text="Condição").grid(row=8, column=0, padx=10, pady=10)
condicao.grid(row=8, column=1, padx=10, pady=10)


# Tipo de anuncio
select_anun = ['Premium', 'Clássico']
tipo_anuncio = tb.Combobox(root,  bootstyle="primary", values=select_anun)
tipo_anuncio.set("Selecione uma opção")
tb.Label(root, text="Tipo Anúncio").grid(row=9, column=0, padx=10, pady=10)
tipo_anuncio.grid(row=9, column=1, padx=10, pady=10)


# Forma envio
select_envio = ['Mercado Envios']
forma_envio = tb.Combobox(root,  bootstyle="primary", values=select_envio)
forma_envio.set("Selecione uma opção")
tb.Label(root, text="Tipo envio").grid(row=10, column=0, padx=10, pady=10)
forma_envio.grid(row=10, column=1, padx=10, pady=10)

# Frete
select_frete = ['Por conta do comprador', 'Você oferece frete grátis']
frete = tb.Combobox(root,  bootstyle="primary", values=select_frete)
frete.set("Selecione uma opção")
tb.Label(root, text="Frete").grid(row=11, column=0, padx=10, pady=10)
frete.grid(row=11, column=1, padx=10, pady=10)

varia_frete = BooleanVar()
check_frete = tb.Checkbutton(bootstyle="primary",
                              text="Variar o frete?",
                              variable=varia_frete,
                              onvalue=True,
                              offvalue=False)

check_frete.grid(row=11, column=2, padx=10, pady=10)

# Condiçao de retirada
select_retirada = ['Concordo', 'Não aceito']
retirada = tb.Combobox(root,  bootstyle="primary", values=select_retirada)
retirada.set("Selecione uma opção")
tb.Label(root, text="Retirada").grid(row=12, column=0, padx=10, pady=10)
retirada.grid(row=12, column=1, padx=10, pady=10)


button = tb.Button(root, text="Abrir Planilha ML", bootstyle="primary outline", command=open_file)
button.grid(row=13, column=0, padx=10, pady=10 )

button = tb.Button(root, text="Abrir Planilha MODELOS", bootstyle="primary outline", command=open_file_models)
button.grid(row=13, column=1, padx=10, pady=10 )

button = tb.Button(root, text="Adicionar dados", bootstyle="warning outline", command=add_data)
button.grid(row=13, column=2, padx=10, pady=10 )

button = tb.Button(root, text="Salvar planilha", bootstyle="danger outline", command=save_file)
button.grid(row=13, column=3, padx=10, pady=10 )



root.mainloop()